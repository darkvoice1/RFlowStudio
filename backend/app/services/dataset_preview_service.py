from csv import DictReader
from csv import Error as CsvError
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook

from app.core.exceptions import DatasetPreviewError
from app.schemas.dataset import (
    DatasetColumnProfile,
    DatasetPreviewResponse,
    DatasetProfileResponse,
    DatasetRecord,
)


class DatasetPreviewService:
    """封装数据集预览与字段分析逻辑。"""

    def get_dataset_preview(
        self,
        record: DatasetRecord,
        data_file_path: Path,
        limit: int,
    ) -> DatasetPreviewResponse:
        """按数据集记录返回当前支持格式的预览结果。"""
        columns, rows, has_more = self._read_preview_data(record, data_file_path, limit)

        return DatasetPreviewResponse(
            dataset_id=record.id,
            file_name=record.file_name,
            columns=columns,
            rows=rows,
            preview_row_count=len(rows),
            limit=limit,
            has_more=has_more,
            preview_format=record.extension.lstrip("."),
        )

    def get_dataset_profile(
        self,
        record: DatasetRecord,
        data_file_path: Path,
    ) -> DatasetProfileResponse:
        """按数据集记录返回当前支持格式的字段元信息统计结果。"""
        columns, column_values, row_count = self._read_profile_data(record, data_file_path)
        profiles = [
            self._build_column_profile(name=column, values=column_values[column])
            for column in columns
        ]

        return DatasetProfileResponse(
            dataset_id=record.id,
            file_name=record.file_name,
            row_count=row_count,
            column_count=len(columns),
            columns=profiles,
            profile_format=record.extension.lstrip("."),
        )

    def _read_preview_data(
        self,
        record: DatasetRecord,
        data_file_path: Path,
        limit: int,
    ) -> tuple[list[str], list[dict[str, str | None]], bool]:
        """按文件格式读取预览数据。"""
        if record.extension == ".csv":
            return self._read_csv_preview(data_file_path=data_file_path, limit=limit)
        if record.extension == ".xlsx":
            return self._read_xlsx_preview(data_file_path=data_file_path, limit=limit)

        raise DatasetPreviewError("当前预览接口暂不支持该文件格式。")

    def _read_profile_data(
        self,
        record: DatasetRecord,
        data_file_path: Path,
    ) -> tuple[list[str], dict[str, list[str | None]], int]:
        """按文件格式读取字段分析所需的列值。"""
        if record.extension == ".csv":
            return self._read_csv_profile(data_file_path=data_file_path)
        if record.extension == ".xlsx":
            return self._read_xlsx_profile(data_file_path=data_file_path)

        raise DatasetPreviewError("当前字段分析接口暂不支持该文件格式。")

    def _read_csv_preview(
        self,
        data_file_path: Path,
        limit: int,
    ) -> tuple[list[str], list[dict[str, str | None]], bool]:
        """读取 CSV 预览数据，返回列名、行数据和是否还有更多行。"""
        try:
            with data_file_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
                reader = DictReader(file_obj)
                columns = reader.fieldnames or []
                if not columns:
                    raise DatasetPreviewError("CSV 文件缺少表头，暂时无法预览。")

                rows: list[dict[str, str | None]] = []
                has_more = False

                # 只读取前 limit 行，避免预览接口一次性把大文件全部拉进内存。
                for row in reader:
                    if len(rows) >= limit:
                        has_more = True
                        break

                    cleaned_row = {column: row.get(column) for column in columns}
                    rows.append(cleaned_row)
        except CsvError as exc:
            raise DatasetPreviewError("CSV 文件格式异常，暂时无法预览。") from exc

        return columns, rows, has_more

    def _read_xlsx_preview(
        self,
        data_file_path: Path,
        limit: int,
    ) -> tuple[list[str], list[dict[str, str | None]], bool]:
        """读取 XLSX 预览数据，返回列名、行数据和是否还有更多行。"""
        try:
            workbook = load_workbook(filename=data_file_path, read_only=True, data_only=True)
        except BadZipFile as exc:
            raise DatasetPreviewError("XLSX 文件格式异常，暂时无法预览。") from exc

        try:
            worksheet = workbook.active
            row_iterator = worksheet.iter_rows(values_only=True)
            header_row = next(row_iterator, None)
            columns = self._extract_header_columns(
                header_row=header_row,
                empty_message="XLSX 文件缺少表头，暂时无法预览。",
            )

            rows: list[dict[str, str | None]] = []
            has_more = False

            # 只保留前 limit 行，避免大表格预览一次性读取过多内容。
            for row in row_iterator:
                if len(rows) >= limit:
                    has_more = True
                    break

                rows.append(self._build_row_dict(columns=columns, values=row))
        finally:
            workbook.close()

        return columns, rows, has_more

    def _read_csv_profile(
        self,
        data_file_path: Path,
    ) -> tuple[list[str], dict[str, list[str | None]], int]:
        """读取 CSV 全量列值，用于字段元信息分析。"""
        try:
            with data_file_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
                reader = DictReader(file_obj)
                columns = reader.fieldnames or []
                if not columns:
                    raise DatasetPreviewError("CSV 文件缺少表头，暂时无法分析字段信息。")

                column_values: dict[str, list[str | None]] = {column: [] for column in columns}
                row_count = 0

                # 逐行累积列值，为字段类型和缺失值统计提供输入。
                for row in reader:
                    row_count += 1
                    for column in columns:
                        column_values[column].append(self._normalize_cell_value(row.get(column)))
        except CsvError as exc:
            raise DatasetPreviewError("CSV 文件格式异常，暂时无法分析字段信息。") from exc

        return columns, column_values, row_count

    def _read_xlsx_profile(
        self,
        data_file_path: Path,
    ) -> tuple[list[str], dict[str, list[str | None]], int]:
        """读取 XLSX 全量列值，用于字段元信息分析。"""
        try:
            workbook = load_workbook(filename=data_file_path, read_only=True, data_only=True)
        except BadZipFile as exc:
            raise DatasetPreviewError("XLSX 文件格式异常，暂时无法分析字段信息。") from exc

        try:
            worksheet = workbook.active
            row_iterator = worksheet.iter_rows(values_only=True)
            header_row = next(row_iterator, None)
            columns = self._extract_header_columns(
                header_row=header_row,
                empty_message="XLSX 文件缺少表头，暂时无法分析字段信息。",
            )

            column_values: dict[str, list[str | None]] = {column: [] for column in columns}
            row_count = 0

            # 逐行累积列值，为字段类型和缺失值统计提供输入。
            for row in row_iterator:
                row_count += 1
                row_dict = self._build_row_dict(columns=columns, values=row)
                for column in columns:
                    column_values[column].append(row_dict[column])
        finally:
            workbook.close()

        return columns, column_values, row_count

    def _extract_header_columns(
        self,
        header_row: tuple[Any, ...] | None,
        empty_message: str,
    ) -> list[str]:
        """从表头行中提取列名列表。"""
        if header_row is None:
            raise DatasetPreviewError(empty_message)

        columns = [self._stringify_header_cell(value) for value in header_row]
        if not any(columns):
            raise DatasetPreviewError(empty_message)

        return columns

    def _stringify_header_cell(self, value: Any) -> str:
        """把表头单元格转换成可展示的列名文本。"""
        if value is None:
            return ""

        return str(value).strip()

    def _build_row_dict(
        self,
        columns: list[str],
        values: tuple[Any, ...] | None,
    ) -> dict[str, str | None]:
        """根据列名和行值构造统一的行字典。"""
        if values is None:
            values = tuple()

        row_dict: dict[str, str | None] = {}
        for index, column in enumerate(columns):
            cell_value = values[index] if index < len(values) else None
            row_dict[column] = self._normalize_cell_value(cell_value)

        return row_dict

    def _build_column_profile(
        self,
        name: str,
        values: list[str | None],
    ) -> DatasetColumnProfile:
        """根据列值列表构造字段元信息。"""
        normalized_values = [self._normalize_cell_value(value) for value in values]
        non_empty_values = [value for value in normalized_values if value is not None]
        inferred_type = self._infer_column_type(non_empty_values)

        # 只保留前几个非空样例，方便前端快速展示字段内容特征。
        sample_values: list[str] = []
        for value in non_empty_values:
            if value not in sample_values:
                sample_values.append(value)
            if len(sample_values) >= 3:
                break

        return DatasetColumnProfile(
            name=name,
            inferred_type=inferred_type,
            nullable=len(non_empty_values) != len(normalized_values),
            missing_count=len(normalized_values) - len(non_empty_values),
            unique_count=len(set(non_empty_values)),
            sample_values=sample_values,
        )

    def _normalize_cell_value(self, value: Any) -> str | None:
        """把原始单元格值归一化为空值或去除首尾空格的文本。"""
        if value is None:
            return None

        normalized = str(value).strip()
        return normalized or None

    def _infer_column_type(
        self,
        values: list[str],
    ) -> str:
        """根据当前列的非空值推断字段类型。"""
        if not values:
            return "empty"
        if all(self._is_integer(value) for value in values):
            return "integer"
        if all(self._is_float(value) for value in values):
            return "float"
        if all(self._is_boolean(value) for value in values):
            return "boolean"
        return "string"

    def _is_integer(self, value: str) -> bool:
        """判断字符串是否可以解析为整数。"""
        try:
            int(value)
        except ValueError:
            return False

        return True

    def _is_float(self, value: str) -> bool:
        """判断字符串是否可以解析为浮点数。"""
        try:
            float(value)
        except ValueError:
            return False

        return True

    def _is_boolean(self, value: str) -> bool:
        """判断字符串是否属于常见布尔值文本。"""
        return value.lower() in {"true", "false", "yes", "no", "0", "1"}
