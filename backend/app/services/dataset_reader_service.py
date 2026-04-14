from csv import DictReader
from csv import Error as CsvError
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook

from app.core.exceptions import DatasetPreviewError
from app.schemas.dataset import DatasetRecord


class DatasetReaderService:
    """负责按文件格式读取数据集行数据。"""

    def read_all_rows(
        self,
        record: DatasetRecord,
        data_file_path: Path,
        empty_message: str,
        csv_header_message: str,
        xlsx_header_message: str,
        xlsx_invalid_message: str,
    ) -> tuple[list[str], list[dict[str, str | None]]]:
        """按文件格式读取全部行数据，供预览和字段分析复用。"""
        if record.extension == ".csv":
            return self._read_csv_rows(
                data_file_path=data_file_path,
                csv_header_message=csv_header_message,
            )
        if record.extension == ".xlsx":
            return self._read_xlsx_rows(
                data_file_path=data_file_path,
                xlsx_header_message=xlsx_header_message,
                xlsx_invalid_message=xlsx_invalid_message,
            )

        raise DatasetPreviewError(empty_message)

    def _read_csv_rows(
        self,
        data_file_path: Path,
        csv_header_message: str,
    ) -> tuple[list[str], list[dict[str, str | None]]]:
        """读取 CSV 全量行数据。"""
        try:
            with data_file_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
                reader = DictReader(file_obj)
                columns = reader.fieldnames or []
                if not columns:
                    raise DatasetPreviewError(csv_header_message)

                rows: list[dict[str, str | None]] = []
                for row in reader:
                    cleaned_row = {
                        column: self._normalize_cell_value(row.get(column))
                        for column in columns
                    }
                    rows.append(cleaned_row)
        except CsvError as exc:
            raise DatasetPreviewError("CSV 文件格式异常，暂时无法预览。") from exc

        return columns, rows

    def _read_xlsx_rows(
        self,
        data_file_path: Path,
        xlsx_header_message: str,
        xlsx_invalid_message: str,
    ) -> tuple[list[str], list[dict[str, str | None]]]:
        """读取 XLSX 全量行数据。"""
        try:
            workbook = load_workbook(filename=data_file_path, read_only=True, data_only=True)
        except BadZipFile as exc:
            raise DatasetPreviewError(xlsx_invalid_message) from exc

        try:
            worksheet = workbook.active
            row_iterator = worksheet.iter_rows(values_only=True)
            header_row = next(row_iterator, None)
            columns = self._extract_header_columns(
                header_row=header_row,
                empty_message=xlsx_header_message,
            )

            rows: list[dict[str, str | None]] = []
            for row in row_iterator:
                rows.append(self._build_row_dict(columns=columns, values=row))
        finally:
            workbook.close()

        return columns, rows

    def _extract_header_columns(
        self,
        header_row: tuple[Any, ...] | None,
        empty_message: str,
    ) -> list[str]:
        """从表头行中提取列名列表。"""
        if header_row is None:
            raise DatasetPreviewError(empty_message)

        columns = [self._stringify_header_cell(value) for value in header_row]
        if not any(columns):
            raise DatasetPreviewError(empty_message)

        return columns

    def _stringify_header_cell(self, value: Any) -> str:
        """把表头单元格转换成可展示的列名文本。"""
        if value is None:
            return ""

        return str(value).strip()

    def _build_row_dict(
        self,
        columns: list[str],
        values: tuple[Any, ...] | None,
    ) -> dict[str, str | None]:
        """根据列名和行值构造统一的行字典。"""
        if values is None:
            values = tuple()

        row_dict: dict[str, str | None] = {}
        for index, column in enumerate(columns):
            cell_value = values[index] if index < len(values) else None
            row_dict[column] = self._normalize_cell_value(cell_value)

        return row_dict

    def _normalize_cell_value(self, value: Any) -> str | None:
        """把原始单元格值归一化为空值或去除首尾空格的文本。"""
        if value is None:
            return None

        normalized = str(value).strip()
        return normalized or None
